import openpyxl
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook('/Users/kiminori/semi-retire-app/Semi-Retire Simulator - 20240531_YT.xlsx')
ws = wb['V4.2']

# Part 2: Focus on areas that were truncated
# Input form area (B4-E36)
print("=" * 120)
print("### INPUT FORM AREA — ROWS 4-36, COLUMNS B-H ###")
print("=" * 120)
for row in range(4, 37):
    for col_idx in range(2, 9):  # B-H
        cell = ws.cell(row=row, column=col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.value is not None:
            print(f"  {col_letter}{row} = {repr(cell.value)}")

# One-time events area (rows 37-65)
print("\n" + "=" * 120)
print("### ONE-TIME EVENTS INPUT AREA — ROWS 37-65, COLUMNS B-H ###")
print("=" * 120)
for row in range(37, 66):
    for col_idx in range(2, 9):  # B-H
        cell = ws.cell(row=row, column=col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.value is not None:
            print(f"  {col_letter}{row} = {repr(cell.value)}")

# Income details input area (rows 66-105)
print("\n" + "=" * 120)
print("### INCOME DETAILS INPUT AREA — ROWS 66-105, COLUMNS B-H ###")
print("=" * 120)
for row in range(66, 106):
    for col_idx in range(2, 9):  # B-H
        cell = ws.cell(row=row, column=col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.value is not None:
            print(f"  {col_letter}{row} = {repr(cell.value)}")

# Column V (one-time events) for rows 4-60
print("\n" + "=" * 120)
print("### COLUMN V (一時収支) — ROWS 4 to 60 ###")
print("=" * 120)
col_idx = openpyxl.utils.column_index_from_string('V')
for row in range(4, 61):
    cell = ws.cell(row=row, column=col_idx)
    if cell.value is not None:
        print(f"  V{row} = {repr(cell.value)}")

# Column AB (Income Reduce rate) - rows 4-60
print("\n" + "=" * 120)
print("### COLUMN AB (Income Reduce rate) — ROWS 4 to 60 ###")
print("=" * 120)
col_idx = openpyxl.utils.column_index_from_string('AB')
for row in range(4, 61):
    cell = ws.cell(row=row, column=col_idx)
    if cell.value is not None:
        print(f"  AB{row} = {repr(cell.value)}")

# Columns W-AA (retirement income #1-#5) - rows 4-10 (pattern)
print("\n" + "=" * 120)
print("### COLUMNS W-AA (Retirement Income #1-#5) — ROWS 4 to 10 ###")
print("=" * 120)
for col_letter in ['W', 'X', 'Y', 'Z', 'AA']:
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    print(f"\n  COLUMN {col_letter}:")
    for row in range(4, 11):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            print(f"    {col_letter}{row} = {repr(cell.value)}")

# Columns AC-AG (retirement income at retire time) - rows 4-10 (pattern)
print("\n" + "=" * 120)
print("### COLUMNS AC-AG (Retirement Income at retire time) — ROWS 4 to 10 ###")
print("=" * 120)
for col_letter in ['AC', 'AD', 'AE', 'AF', 'AG']:
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    print(f"\n  COLUMN {col_letter}:")
    for row in range(4, 11):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            print(f"    {col_letter}{row} = {repr(cell.value)}")

# Column AH (death estate) - rows 4-60
print("\n" + "=" * 120)
print("### COLUMN AH (死亡時の遺産) — ROWS 4 to 60 ###")
print("=" * 120)
col_idx = openpyxl.utils.column_index_from_string('AH')
for row in range(4, 61):
    cell = ws.cell(row=row, column=col_idx)
    if cell.value is not None:
        print(f"  AH{row} = {repr(cell.value)}")

# Columns AJ-AK (present value calc) - rows 4-10
print("\n" + "=" * 120)
print("### COLUMNS AJ-AK (Present Value Calc) — ROWS 4 to 10 ###")
print("=" * 120)
for col_letter in ['AJ', 'AK']:
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    print(f"\n  COLUMN {col_letter}:")
    for row in range(4, 11):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            print(f"    {col_letter}{row} = {repr(cell.value)}")

# Result summary area
print("\n" + "=" * 120)
print("### RESULT SUMMARY — ROWS 100-120, ALL COLUMNS A-AK ###")
print("=" * 120)
for row in range(100, 121):
    for col_idx in range(1, 38):
        cell = ws.cell(row=row, column=col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.value is not None:
            print(f"  {col_letter}{row} = {repr(cell.value)}")

# Column S (year-end balance) - rows 4-60 pattern check
print("\n" + "=" * 120)
print("### COLUMN S (年末残高) — ROWS 4 to 15 ###")
print("=" * 120)
col_idx = openpyxl.utils.column_index_from_string('S')
for row in range(4, 16):
    cell = ws.cell(row=row, column=col_idx)
    if cell.value is not None:
        print(f"  S{row} = {repr(cell.value)}")

# Column Q (adjustment) - rows 4-15
print("\n" + "=" * 120)
print("### COLUMN Q (調整額) — ROWS 4 to 15 ###")
print("=" * 120)
col_idx = openpyxl.utils.column_index_from_string('Q')
for row in range(4, 16):
    cell = ws.cell(row=row, column=col_idx)
    if cell.value is not None:
        print(f"  Q{row} = {repr(cell.value)}")

# Column R (living expenses pre-tax) - rows 4-15
print("\n" + "=" * 120)
print("### COLUMN R (生活費 税引前) — ROWS 4 to 15 ###")
print("=" * 120)
col_idx = openpyxl.utils.column_index_from_string('R')
for row in range(4, 16):
    cell = ws.cell(row=row, column=col_idx)
    if cell.value is not None:
        print(f"  R{row} = {repr(cell.value)}")

# Column M (pre-retirement annual income) - rows 4-15
print("\n" + "=" * 120)
print("### COLUMN M (リタイア迄の年間収支) — ROWS 4 to 15 ###")
print("=" * 120)
col_idx = openpyxl.utils.column_index_from_string('M')
for row in range(4, 16):
    cell = ws.cell(row=row, column=col_idx)
    if cell.value is not None:
        print(f"  M{row} = {repr(cell.value)}")

print("\n\nDone.")

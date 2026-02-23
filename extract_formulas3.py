import openpyxl
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook('/Users/kiminori/semi-retire-app/Semi-Retire Simulator - 20240531_YT.xlsx')
ws = wb['V4.2']

with open('/Users/kiminori/semi-retire-app/formulas_output.txt', 'w') as f:
    # INPUT FORM
    f.write("=" * 120 + "\n")
    f.write("### INPUT FORM AREA — ROWS 4-36, COLUMNS B-H ###\n")
    f.write("=" * 120 + "\n")
    for row in range(4, 37):
        for col_idx in range(2, 9):
            cell = ws.cell(row=row, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if cell.value is not None:
                f.write(f"  {col_letter}{row} = {repr(cell.value)}\n")

    # ONE-TIME EVENTS + INCOME INPUTS
    f.write("\n" + "=" * 120 + "\n")
    f.write("### ONE-TIME EVENTS + INCOME INPUT AREA — ROWS 37-105, COLUMNS B-H ###\n")
    f.write("=" * 120 + "\n")
    for row in range(37, 106):
        for col_idx in range(2, 9):
            cell = ws.cell(row=row, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if cell.value is not None:
                f.write(f"  {col_letter}{row} = {repr(cell.value)}\n")

    # G column rows 37-65
    f.write("\n" + "=" * 120 + "\n")
    f.write("### G COLUMN (AGE MAPPING for one-time events) — ROWS 37-65 ###\n")
    f.write("=" * 120 + "\n")
    for row in range(37, 66):
        cell = ws.cell(row=row, column=7)
        if cell.value is not None:
            f.write(f"  G{row} = {repr(cell.value)}\n")

    # C column rows 66-105
    f.write("\n" + "=" * 120 + "\n")
    f.write("### C COLUMN — ROWS 66-105 ###\n")
    f.write("=" * 120 + "\n")
    for row in range(66, 106):
        cell = ws.cell(row=row, column=3)
        if cell.value is not None:
            f.write(f"  C{row} = {repr(cell.value)}\n")

    # D column rows 37-105
    f.write("\n" + "=" * 120 + "\n")
    f.write("### D COLUMN — ROWS 37-105 ###\n")
    f.write("=" * 120 + "\n")
    for row in range(37, 106):
        cell = ws.cell(row=row, column=4)
        if cell.value is not None:
            f.write(f"  D{row} = {repr(cell.value)}\n")

    # ALL COLUMNS for ROW 4, 5, 6
    f.write("\n" + "=" * 120 + "\n")
    f.write("### ALL FORMULAS — ROWS 4-6, COLUMNS A-AK ###\n")
    f.write("=" * 120 + "\n")
    for row in [4, 5, 6]:
        f.write(f"\n--- ROW {row} ---\n")
        for col_idx in range(1, 38):
            cell = ws.cell(row=row, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if cell.value is not None:
                f.write(f"  {col_letter}{row} = {repr(cell.value)}\n")

    # KEY COLUMNS pattern (rows 4-15)
    for col_letter in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'V', 'W', 'X', 'AB', 'AC', 'AH', 'AJ', 'AK']:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        header = ws.cell(row=3, column=col_idx).value or ws.cell(row=2, column=col_idx).value or ''
        f.write(f"\n### COLUMN {col_letter} ({header}) — ROWS 4 to 15 ###\n")
        for row in range(4, 16):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                f.write(f"  {col_letter}{row} = {repr(cell.value)}\n")

    # RESULT SUMMARY
    f.write("\n" + "=" * 120 + "\n")
    f.write("### RESULT SUMMARY — ROWS 100-120, ALL COLUMNS ###\n")
    f.write("=" * 120 + "\n")
    for row in range(100, 121):
        for col_idx in range(1, 38):
            cell = ws.cell(row=row, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if cell.value is not None:
                f.write(f"  {col_letter}{row} = {repr(cell.value)}\n")

    # DEFINED NAMES
    f.write("\n" + "=" * 120 + "\n")
    f.write("### DEFINED NAMES ###\n")
    f.write("=" * 120 + "\n")
    try:
        for name, defn in wb.defined_names.items():
            f.write(f"  {name} = {defn.attr_text}\n")
    except Exception as e:
        f.write(f"  (Could not read defined names: {e})\n")

print("Done. Written to /Users/kiminori/semi-retire-app/formulas_output.txt")

import openpyxl
import sys

wb = openpyxl.load_workbook(
    "/Users/kiminori/semi-retire-app/Semi-Retire Simulator - 20240531_YT.xlsx",
    data_only=False
)

ws = wb["V4.2"]

# Columns to extract for rows 4-10
columns = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'W', 'X', 'AC', 'AD', 'AB']

# Additional specific cells
specific_cells = [
    'M4', 'M5', 'M6', 'M7',
    'W4', 'W5', 'W6',
    'AC4', 'AC5', 'AC6',
    'AB4', 'AB5',
    'AJ4', 'AJ5',
    'AK4', 'AK5',
]

output_lines = []

def log(msg=""):
    print(msg)
    output_lines.append(msg)

log("=" * 80)
log("FORMULA EXTRACTION from V4.2 sheet")
log("=" * 80)

# Part 1: Rows 4-10, specified columns
log("")
log("PART 1: Rows 4-10 for columns M, N, O, P, Q, R, S, W, X, AC, AD, AB")
log("-" * 80)

for row in range(4, 11):
    log(f"\n--- Row {row} ---")
    for col in columns:
        cell_ref = f"{col}{row}"
        cell = ws[cell_ref]
        val = cell.value
        if val is None:
            display = "<empty>"
        elif isinstance(val, str) and val.startswith("="):
            display = val
        else:
            display = repr(val)
        log(f"  {cell_ref:>5s} = {display}")

# Part 2: Specific cells highlighted
log("")
log("=" * 80)
log("PART 2: Specific cells of interest")
log("=" * 80)

groups = [
    ("M4-M7 (checking if M6 differs from M5)", ['M4', 'M5', 'M6', 'M7']),
    ("W4-W6", ['W4', 'W5', 'W6']),
    ("AC4-AC6", ['AC4', 'AC5', 'AC6']),
    ("AB4-AB5", ['AB4', 'AB5']),
    ("AJ4-AJ5", ['AJ4', 'AJ5']),
    ("AK4-AK5", ['AK4', 'AK5']),
]

for group_name, cells in groups:
    log(f"\n--- {group_name} ---")
    for cell_ref in cells:
        cell = ws[cell_ref]
        val = cell.value
        if val is None:
            display = "<empty>"
        elif isinstance(val, str) and val.startswith("="):
            display = val
        else:
            display = repr(val)
        log(f"  {cell_ref:>5s} = {display}")

# Save to file
with open("/Users/kiminori/semi-retire-app/formula_details.txt", "w") as f:
    f.write("\n".join(output_lines) + "\n")

log("")
log("Results saved to /Users/kiminori/semi-retire-app/formula_details.txt")
